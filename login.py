from tkinter import *
from openpyxl import *
import pandas as pd

wb=load_workbook("C:\\Users\\Abhayraj sinh parmar\\python_files\\Exam_dashboard\\Excel Sheets\\student_info.xlsx")

sheet=wb.active

def clear():
    email_id_field.delete(0, END)
    password_field.delete(0, END)

def insert():

    if (email_id_field.get()=="" and password_field.get()==""):
        print("Empty input")
    elif (email_id_field.get()==""):
        print("Mail should not be empty")
    elif (password_field.get()==""):
        print("Password should not be  empty")
    else:      
        print("Mail and password do not match")

    for row in sheet.iter_rows(values_only=True):
        if (row[1] == email_id_field.get() and row[2] == password_field.get()):
            print("login succesful")
            email_id = email_id_field.get()
            clear()
            root.destroy()
            display_questions(email_id)
            break


def display_questions(email_id):
    questions_df = pd.read_excel("C:\\Users\\Abhayraj sinh parmar\\python_files\\Exam_dashboard\\Excel Sheets\\Questions.xlsx",usecols='A:F')
    answers_df = pd.read_excel("C:\\Users\\Abhayraj sinh parmar\\python_files\\Exam_dashboard\\Excel Sheets\\Answers.xlsx", header=None, index_col=0)
    
    score = 0
    
    for index, row in questions_df.iterrows():
        question_key = row[0]
        question_text = row[1]
        options = row[2:6]  
        print(f"Question {index + 1}: {question_text}") 

        for  i in range(2):
            for j in range(2):
                option_index= i*2 + j  
                if option_index<len(options):
                    print(f"{option_index + 1} {options[option_index]}",end="\t") 
            print()

        user_answer = input("Your answer (1-4): ")
 
        try:
            user_choice = int(user_answer)
            if 1 <= user_choice <= 4:
                correct_answer_index = answers_df.loc[question_key, 1] - 1  
                correct_answer = options[correct_answer_index]
                if options[user_choice - 1] == correct_answer:
                    print("Correct!\n")
                    score += 1
                else:
                    print(f"Incorrect. The correct answer is: {correct_answer}\n")
            else:
                print("Invalid input. Please enter a number between 1 and 4.\n")
        except ValueError:
            print("Invalid input. Please enter a number between 1 and 4.\n")
        except KeyError:
            print(f"KeyError: The key '{question_key}' was not found in answers.xlsx")


    print(f"Your final score is: {score}/{len(questions_df)}")
    store_result(email_id, score, len(questions_df))

def store_result(email_id, score, total_questions):
    result_file = "C:\\Users\\Abhayraj sinh parmar\\python_files\\Exam_dashboard\\Excel Sheets\\Result.xlsx"
    
    result_wb = load_workbook(result_file)
    result_sheet = result_wb.active
    
    result_sheet.append([email_id, score, total_questions])
    result_wb.save(result_file)
    print("Results saved successfully.")
    

def main():
    global root, email_id_field, password_field
    root=Tk()
    
    root.config(background="light blue")
    root.geometry("450x175")
    root.title("Login Form")

    heading=Label(root, text="Form", bg="light blue")
    email=Label(root, text="email", bg="light blue")
    password=Label(root, text="password", bg="light blue")  

    heading.grid(row=0, column=1)
    email.grid(row=2, column=0)
    password.grid(row=3, column=0)

    
    email_id_field=Entry(root)
    password_field=Entry(root, show="*")


    email_id_field.grid(row=2, column=1, ipadx="100")
    password_field.grid(row=3, column=1, ipadx="100")   

        #  excel()

    email_id_field.grid(row=2, column=1, ipadx="100")
    password_field.grid(row=3, column=1, ipadx="100")

    submit=Button(root, text="submit", command=insert)

    submit.grid(row=6,column=1)

    root.mainloop()



try:
    if __name__ == "__main__":
        main()
except:
    print("Callback terminated...")
       