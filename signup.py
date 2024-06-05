from openpyxl import *
from tkinter import *
import re
import login

wb=load_workbook("C:\\Users\\Abhayraj sinh parmar\\python_files\\Exam_dashboard\\Excel Sheets\\Student_info.xlsx")

sheet=wb.active 

def excel():
    sheet.column_dimensions['A'].width=30
    sheet.column_dimensions['B'].width=40
    sheet.column_dimensions['C'].width=18
    sheet.column_dimensions['D'].width=18
    sheet.column_dimensions['E'].width=10

    sheet.cell(row=1, column=1).value="Username"
    sheet.cell(row=1, column=2).value="Email id"
    sheet.cell(row=1, column=3).value="Password"
    sheet.cell(row=1, column=4).value="confirmPassword"
    sheet.cell(row=1, column=5).value="D.O.B"

def focus1(event):
    email_id_field.focus_set()

def focus2(event):
    password_field.focus_set()

def focus3(event):
    confirm_password_field.focus_set()

def focus4(event):
    date_of_birth_field.focus_set() 



def clear():
     username_field.delete(0, END)
     email_id_field.delete(0, END)
     password_field.delete(0, END)
     confirm_password_field.delete(0, END)
     date_of_birth_field.delete(0, END)

def insert():
    if (username_field.get() == "" and
       email_id_field.get() == "" and
       password_field.get() == "" and
       confirm_password_field.get()== "" and
       date_of_birth_field.get() == ""):
        
        print("Empty input")
    elif (username_field.get()==""):
        print("Username should not be empty")
    elif (email_id_field.get()==""):
        print("Email should not be empty")
    elif (password_field.get()==""):
        print("passoword should not be empty")
    elif (confirm_password_field.get()==""):
        print("passoword should not be empty")
    elif (date_of_birth_field.get()==""):
        print("date of birth should not be empty")

    elif not (email_id_field.get().endswith("@gmail.com") ) or (email_id_field.get().startswith(" ")) or email_id_field.get().startswith("@gmail.com"):     
        print("Mail is not in recognized")

    elif (password_field.get() != confirm_password_field.get()):
        print("Password do not match")

    elif not re.match(r'^\d{2}-\d{2}-\d{4}$', date_of_birth_field.get()):
        print("Date of Birth format should be dd-mm-yyyy")

    else:
        
        current_row=sheet.max_row
        current_column=sheet.max_column

        sheet.cell(row=current_row+1, column=1).value=username_field.get()
        sheet.cell(row=current_row+1, column=2).value=email_id_field.get()
        sheet.cell(row=current_row+1, column=3).value=password_field.get()
        sheet.cell(row=current_row+1, column=4).value=confirm_password_field.get()
        sheet.cell(row=current_row+1, column=5).value=date_of_birth_field.get()

        wb.save("C:\\Users\\Abhayraj sinh parmar\\python_files\\Exam_dashboard\\Excel Sheets\\Student_info.xlsx")
       
        username_field.focus_set()

        clear()
        root.destroy()
        
def log():
    root.destroy()
    login.main()

if __name__ == "__main__":

    root=Tk()
    
    root.config(background="light blue")
    root.geometry("500x250")
    root.title("Signup Form")

    excel()

    heading=Label(root, text="Form", bg="light blue")
    username=Label(root, text="Username", bg="light blue")
    email=Label(root, text="email", bg="light blue")
    password=Label(root, text="password", bg="light blue")
    confirm_pass=Label(root, text="Confirm Password", bg="light blue")
    date_of_birth=Label(root, text="Date of Birth", bg="light blue")

    heading.grid(row=0, column=1)
    username.grid(row=1, column=0)
    email.grid(row=2, column=0)
    password.grid(row=3, column=0)
    confirm_pass.grid(row=4, column=0)
    date_of_birth.grid(row=5, column=0)

    username_field=Entry(root)
    email_id_field=Entry(root)
    password_field=Entry(root, show="*")
    confirm_password_field=Entry(root, show="*")
    date_of_birth_field=Entry(root)
    username_field.bind("<Return>",focus1)
    email_id_field.bind("<Return>", focus2)
    password_field.bind("<Return>", focus3)
    confirm_password_field.bind("<Return>", focus4)
    
    

    username_field.grid(row=1, column=1, ipadx="100")
    email_id_field.grid(row=2, column=1, ipadx="100")
    password_field.grid(row=3, column=1, ipadx="100")
    confirm_password_field.grid(row=4, column=1, ipadx="100")
    date_of_birth_field.grid(row=5, column=1, ipadx="100")

    excel()

    submit= Button(root,text="submit", command=insert)
    loginbtn= Button(root, text="Login", command=log)
    submit.grid(row=6,column=1)
    loginbtn.grid(row=7,column=1)

    root.mainloop()